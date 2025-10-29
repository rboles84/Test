"""Loader for spreadsheet artifacts (CSV and XLSX)."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from rag.ingestion.base_loader import ArtifactLoader
from rag.models import ArtifactRecord


class SpreadsheetLoader(ArtifactLoader):
    """Extracts rows from spreadsheets as plain text."""

    def __init__(self, path: Path) -> None:
        super().__init__(path, doc_type="spreadsheet")

    def _load(self) -> Iterable[ArtifactRecord]:
        suffix = self.path.suffix.lower()
        if suffix == ".csv":
            yield from self._load_csv()
        elif suffix in {".xlsx", ".xlsm"}:
            yield from self._load_xlsx()
        else:
            raise ValueError(f"Unsupported spreadsheet format: {self.path}")

    def _load_csv(self) -> Iterable[ArtifactRecord]:
        with self.path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for idx, row in enumerate(reader, start=1):
                text = " | ".join(f"{key}: {value}" for key, value in row.items())
                yield ArtifactRecord(text=text, metadata={"row": str(idx)})

    def _load_xlsx(self) -> Iterable[ArtifactRecord]:
        workbook = load_workbook(filename=self.path, data_only=True)
        for sheet in workbook.worksheets:
            headers = [cell.value or "" for cell in next(sheet.iter_rows(max_row=1))]
            for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                values = [cell.value for cell in row]
                pairs = [f"{header}: {value}" for header, value in zip(headers, values)]
                text = " | ".join(pairs)
                yield ArtifactRecord(
                    text=text,
                    metadata={"sheet": sheet.title, "row": str(row_index)},
                )
